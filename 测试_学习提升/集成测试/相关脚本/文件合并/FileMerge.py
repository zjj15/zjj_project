import os
import openpyxl

currentRow =12
#判断Facets_Test_Result_RoundOne_en文件是否存在

if os.path.exists('Facets_Test_Result_RoundOne_en.xlsx'):
    
    mywbResult =openpyxl.load_workbook(filename='./Facets_Test_Result_RoundOne_en.xlsx')     
    sheetResults = mywbResult.get_sheet_by_name("FacetsTestPlan")
    maxRow = sheetResults.max_row
    for i in range(currentRow,maxRow+1):
        sheetResults.cell(i,2).value = ""
        sheetResults.cell(i,3).value = ""
        sheetResults.cell(i,5).value = "" 
else:
    error("Facets_Test_Result_RoundOne_en or FacetsTestPlan文件不存在")
    end()
    
    

# Requirements文件处理  
if os.path.exists('Requirements.xlsx'):
    mywbRequirements =openpyxl.load_workbook(filename='./Requirements.xlsx')  
    sheets = mywbRequirements.sheetnames    
    
    for i in range(len(sheets)):
        maxRow = 0
        sheet = mywbRequirements[sheets[i]]
        if sheet['A2'].value=="Test Case ID":
            maxRow = sheet.max_row-1
            for m in range(1,maxRow):
                #copy
                sheetResults.cell(currentRow,2).value = "Requirements"
                sheetResults.cell(currentRow,3).value = sheet.cell(m+2,1).value
                sheetResults.cell(currentRow,5).value = sheet.cell(m+2,6).value
                currentRow = currentRow + 1                                           
else:
    print("Requirements文件不存在")
    
    
# User Scenarios文件处理  
if os.path.exists('User Scenarios.xlsx'):
    mywbRequirements =openpyxl.load_workbook(filename='./User Scenarios.xlsx')  
    sheets = mywbRequirements.sheetnames    
    
    for i in range(len(sheets)):
        maxRow = 0
        sheet = mywbRequirements[sheets[i]]
        if sheet['A2'].value=="Test Case ID":
            maxRow = sheet.max_row-1
            for m in range(1,maxRow):
                #copy
                sheetResults.cell(currentRow,2).value = "User Scenarios"
                sheetResults.cell(currentRow,3).value = sheet.cell(m+2,1).value
                sheetResults.cell(currentRow,5).value = sheet.cell(m+2,6).value
                currentRow = currentRow + 1                                           
else:
    print("User Scenarios文件不存在")
    
# Wired文件处理  
if os.path.exists('User Scenarios.xlsx'):
    mywbRequirements =openpyxl.load_workbook(filename='./Wired.xlsx')  
    sheets = mywbRequirements.sheetnames    
    
    for i in range(len(sheets)):
        maxRow = 0
        sheet = mywbRequirements[sheets[i]]
        if sheet['A2'].value=="Test Case ID":
            maxRow = sheet.max_row-1
            for m in range(1,maxRow):
                #copy
                sheetResults.cell(currentRow,2).value = "Wired"
                sheetResults.cell(currentRow,3).value = sheet.cell(m+2,1).value
                sheetResults.cell(currentRow,5).value = sheet.cell(m+2,6).value
                currentRow = currentRow + 1                                           
else:
    print("Wired文件不存在")
    
# Wireless文件处理  
if os.path.exists('User Scenarios.xlsx'):
    mywbRequirements =openpyxl.load_workbook(filename='./Wireless.xlsx')  
    sheets = mywbRequirements.sheetnames    
    
    for i in range(len(sheets)):
        maxRow = 0
        sheet = mywbRequirements[sheets[i]]
        if sheet['A2'].value=="Test Case ID":
            maxRow = sheet.max_row-1
            for m in range(1,maxRow):
                #copy
                sheetResults.cell(currentRow,2).value = "Wireless"
                sheetResults.cell(currentRow,3).value = sheet.cell(m+2,1).value
                sheetResults.cell(currentRow,5).value = sheet.cell(m+2,6).value
                currentRow = currentRow + 1                                           
else:
    print("Wireless文件不存在")
  
mywbResult.save('Facets_Test_Result_RoundOne_en.xlsx') 
    
    
    
    
    
    
    
    
    