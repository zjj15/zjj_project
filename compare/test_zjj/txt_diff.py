#coding=utf-8
#txt 文档对比

import os
from  openpyxl import Workbook


last_name='lastWeekFiles\installed-files-boot.txt'
new_name='thisWeekFiles\installed-files-boot.txt'
result_name='result_txt.xlsx'

#结果输出文档
wb=Workbook()
#2.写入
ws=wb.active#print(ws)#<Worksheet "Sheet">
'''
ws['A1']='上周Release相对本周多出的文件'
ws['B1']='本周Release相对上周多出的文件'
ws['C1']='大小无变化的文件'
ws['D1']='大小有变化的文件'
ws['E1']='变化量'
'''


#对比文档
f1=open(last_name, mode='r')
f2=open(new_name, mode='r')


def Copystr(str1, _str):
    for cell in ws[_str]:
        cell.value=str1

    
for f1_readline in f1.readlines():
    for f2_readline  in f2.readlines():
        if f1_readline.find(f2_readline)!=-1:
            print('大小无变化的文件 '+ f1_readline)
            Copystr(f1_readline, 'C')
        elif  f1_readline[0:f1_readline.find('  /')] != f2_readline[0:f2_readline.find('  /')]   and  f1_readline[f1_readline.find('  /'):] == f2_readline[f2_readline.find('  /'):] :
            print('大小有变化的文件')
            Copystr(f2_readline, 'D')
           
            print('变化量')
            Copystr(str(int(f2_readline[0:f2_readline.find('  /')] )- int(f1_readline[0:f1_readline.find('  /')])), 'E')
        else:
            print('上周Release相对本周多出的文件')
            Copystr(f1_readline, 'A')  

#遇到的困难： 1.没法从每一列第二行开始填充数据,在excel里没法按顺序填充数据

wb.save(result_name)
f1.close()
f2.close()

#and  f1_readline[f1_readline.find('  /'):] == f2_readline[f2_readline.find('  /'):]