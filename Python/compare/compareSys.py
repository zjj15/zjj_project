import os
import openpyxl



#打开上周/本周release相关文件
f1 =  open('lastInstallFile1.txt', 'r+',encoding='utf-8')
f2 =  open('lastInstallFile2.txt', 'r+',encoding='utf-8')
#结果输出文档
f3 = open('Result2.txt', 'w',encoding='utf-8') 

#mywb = openpyxl.Workbook(Result.xlsx)
mywb =openpyxl.load_workbook(filename='./Result.xlsx')
mywb.create_sheet(title='result_sys',index = 1)
mysheet = mywb.get_sheet_by_name('result_sys')
mysheet['A1'] = "上周Release相对本周多出的文件"
mysheet['B1'] = "本周Release相对上周多出的文件"
mysheet['C1'] = "大小无变化的文件"
mysheet['D1'] = "大小有变化的文件"
mysheet['E1'] = "变化量"
mysheet.column_dimensions['A'].width = 40
mysheet.column_dimensions['B'].width = 40
mysheet.column_dimensions['C'].width = 40
mysheet.column_dimensions['D'].width = 40
mysheet.column_dimensions['E'].width = 8
#定义list
#存放上周release相关文件
lastReleaseFile = []
#存放本周release相关文件
thisReleaseFile = []
#存放本周release需要删除的文件项目
lastNeedRemoveIndex=[]
#存放本周release需要删除的文件项目
thisNeedRemoveIndex=[]
#存放修改本周release相关文件
tmp1File = []
#存放修改上周release相关文件
tmp2File = []
tmpFile = []
row1 = 2
row2 = 2
row3 = 2
row4 = 2
#将文本中内容放在list中
for line1 in f1:
    lastReleaseFile.append(line1)
for line2 in f2:
    thisReleaseFile.append(line2)
    
#确认增减的文件  
#本周与上周比较   
for i in range(0,len(lastReleaseFile)):
    for j in range(0,len(thisReleaseFile)):
        if (thisReleaseFile[j][14:]) == (lastReleaseFile[i][14:]):
            tmp1File.append(thisReleaseFile[j])
            break
    else:
        lastNeedRemoveIndex.append(i)
        print("thisWeek不存在lastWeek拥有的文件",lastReleaseFile[i][14:]) 
        f3.write("thisWeek不存在lastWeek拥有的文件"+lastReleaseFile[i][14:])
        mysheet.cell(row1,1).value = (lastReleaseFile[i][14:])
        row1 = row1+1
#print("last",len(lastReleaseFile))
#上周与本周比较     
for i in range(0,len(thisReleaseFile)):
    for j in range(0,len(lastReleaseFile)):
        if (lastReleaseFile[j][14:]) == (thisReleaseFile[i][14:]):
            tmp2File.append(lastReleaseFile[j])
            break
    else:
        thisNeedRemoveIndex.append(i)
        print("lastWeek不存在thisWeek拥有的文件",thisReleaseFile[i][14:])    
        f3.write("lastWeek不存在thisWeek拥有的文件"+thisReleaseFile[i][14:])
        mysheet.cell(row2,2).value = (thisReleaseFile[i][14:])
        row2 = row2+1        

'''       
#增删：将上周与本周文件新增或删除的部分上次,此时两文件文件数相同

for i in range(0,len(lastNeedRemoveIndex)):
    lastReleaseFile.pop(lastNeedRemoveIndex[i])
    
for i in range(0,len(thisNeedRemoveIndex)):
    thisReleaseFile.pop(thisNeedRemoveIndex[i])
'''
    
#排序：以上周为base,将本周文件进行排序，存放在tmpFile中

for i in range(0,len(tmp1File)):
    #print(baseList[i])
    for j in range(0,len(tmp2File)):
        if (tmp2File[j][14:]) == (tmp1File[i][14:]):
            tmpFile.append(tmp2File[j])
            break
            
f3.write("\n")            
f3.write("文件大小变化如下：")  
f3.write("\n")  
#确认文件大小的改变            
for i in range(0,len(tmp1File)):
    #print(baseList[i],end="")
    if int(tmp1File[i][0:14]) == int(tmpFile[i][0:14]):
        #print("文件大小没变化:",tmp1File[i][14:])
        mysheet.cell(row3,3).value = (tmp1File[i][14:])
        row3 = row3+1  
    else:
        #print("文件大小变化了:",tmp1File[i][14:],"变化量本周相对上周：",int(tmp1File[i][0:14])-int(tmpFile[i][0:14]))
        #print(tmpFile[i][14:],tmpFile[i][0:14])
        #print(tmp1File[i][14:],tmp1File[i][0:14])
        f3.write("此文件大小有变化："+ tmp1File[i][14:])
        variation = int(tmp1File[i][0:14])-int(tmpFile[i][0:14])
        f3.write("      变化量本周相对上周："+ str(variation)) 
        mysheet.cell(row4,4).value = (tmp1File[i][14:])
        mysheet.cell(row4,5).value = (str(variation))
        row4 = row4+1  


mywb.save('Result.xlsx')

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    