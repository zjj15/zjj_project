import os
import openpyxl


#打开上周/本周release相关文件
f1 =  open('lastWeekFiles\installed-files-system.txt', 'r+',encoding='utf-8')
f2 =  open('thisWeekFiles\installed-files-system.txt', 'r+',encoding='utf-8')

mywb =openpyxl.load_workbook(filename='./Result.xlsx')
mywb.create_sheet(title='result_sys',index = 1)
mysheet = mywb.get_sheet_by_name('result_sys')

mysheet['A1'] = "上周Release文件"
mysheet['B1'] = "本周Release文件"
mysheet['C1'] = "本周Release文件大小"
mysheet['D1'] = "变化量"

mysheet.column_dimensions['A'].width = 40
mysheet.column_dimensions['B'].width = 40
mysheet.column_dimensions['C'].width = 20
mysheet.column_dimensions['D'].width = 10
#定义list
#存放上周release相关文件
lastReleaseFile = []
#存放本周release相关文件
thisReleaseFile = []
#存放本周release需要删除的文件项目
lastNeedRemoveIndex=[]
#存放本周release需要删除的文件项目
thisNeedRemoveIndex=[]
#存放修改本周release相关文件
tmp1File = []
#存放修改上周release相关文件
tmp2File = []
tmpFile = []

#将文本中内容放在list中
for line1 in f1:
    lastReleaseFile.append(line1)
for line2 in f2:
    thisReleaseFile.append(line2)
    
#确认增减的文件  
#本周与上周比较   
for i in range(0,len(lastReleaseFile)):
    for j in range(0,len(thisReleaseFile)):
        if (thisReleaseFile[j][14:]) == (lastReleaseFile[i][14:]):
            tmp1File.append(thisReleaseFile[j])
            break
    else:
        
        lastNeedRemoveIndex.append(i)
       
#上周与本周比较  

for i in range(0,len(thisReleaseFile)):
    for j in range(0,len(lastReleaseFile)):
        if (lastReleaseFile[j][14:]) == (thisReleaseFile[i][14:]):
            tmp2File.append(lastReleaseFile[j])
            break
    else:
        thisNeedRemoveIndex.append(i)
        print("lastWeek不存在thisWeek拥有的文件",thisReleaseFile[i][14:])    
            
#增删：将上周与本周文件新增或删除的部分上次,此时两文件文件数相同
   
#排序：以上周为base,将本周文件进行排序，存放在tmpFile中

for i in range(0,len(tmp1File)):
    #print(baseList[i])
    for j in range(0,len(tmp2File)):
        if (tmp2File[j][14:]) == (tmp1File[i][14:]):
            tmpFile.append(tmp2File[j])
            break
lenEqual = len(tmpFile)
            
#将两次release文件相同部分放在excel中
for i in range(0,len(tmp1File)):
    variation = int(tmpFile[i][0:14])-int(tmp1File[i][0:14])
    mysheet.cell(i+2,1).value = (tmp1File[i][14:])
    mysheet.cell(i+2,2).value = (tmpFile[i][14:])
    mysheet.cell(i+2,3).value = (int(tmp1File[i][0:14]))
    mysheet.cell(i+2,4).value = (str(variation))
    
#将两次release文件不相同部分放在excel中    

for i in range(0,len(lastNeedRemoveIndex)):
    mysheet.cell(lenEqual+2+i,1).value = (lastReleaseFile[lastNeedRemoveIndex[i]][14:])
    mysheet.cell(lenEqual+2+i,2).value = ("-")
    mysheet.cell(lenEqual+2+i,4).value = ("del")

for i in range(0,len(thisNeedRemoveIndex)):
    mysheet.cell(lenEqual+2+i,1).value = ("-")
    mysheet.cell(lenEqual+2+i,2).value = (thisReleaseFile[thisNeedRemoveIndex[i]][14:])
    mysheet.cell(lenEqual+2+i,3).value = (int(tmp1File[i][0:14]))
    mysheet.cell(lenEqual+2+i,4).value = ("new")
mywb.save('Result.xlsx')

   
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    