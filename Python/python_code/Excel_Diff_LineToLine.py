#coding=utf-8
# -*- coding: utf-8 -*-
#Excel文档对比脚本
'''
前提：自己复制到Excel里，因此行列都已知晓
思路：读取所有内容，1 in n进行对比/一行行对比（适用于两个文档数据不相同其余相同），不同处标黄
疑问1：如果两个文档的数据不是一行一行对应的，怎么对比？比如：第一个文档的第3行在第二个文档的第10行能找到，
关于疑问1的想法：这种对比的时候应该是把第一个文档的每一行拎出来在第二个文档里全局查找
疑问2:涉及到符号和空格的怎么办？
关于疑问2的想法：一会儿实际操作遇上了再说
'''

#一行行对比

from  openpyxl import Workbook
from  openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

#读取sheet内容
name='//Users//zhangjingjing//desktop//diff.xlsx'
wb=load_workbook(name)
ws=wb['Sheet1']
ws2=wb['Sheet2']

#一行行比较
for row, row2 in zip(ws[1:405],ws2[1:405]):
    for r,r2 in zip(row,row2):
        if (str(r.value)).find(str(r2.value))== -1:
            #sheet1
            row_split1=(str(row).split('.'))[1]
            row_split2=(row_split1.split('>'))[0]
            ft = Font(color=colors.RED) 
            ws[row_split2].font=ft 
            print(row_split2)
            #sheet2
            row2_split1=(str(row2).split('.'))[1]
            row2_split2=(row2_split1.split('>'))[0]
            ft = Font(color=colors.RED)
            ws2[row2_split2].font=ft  
        else:
            pass 
#保存
wb.save(name)
