#coding=utf-8
# -*- coding: utf-8 -*-
#Excel文档对比脚本
'''
前提：自己复制到Excel里，因此行列都已知晓
思路：读取所有内容，1 in n进行对比/一行行对比（适用于两个文档数据不相同其余相同），不同处标黄
疑问1：如果两个文档的数据不是一行一行对应的，怎么对比？比如：第一个文档的第3行在第二个文档的第10行能找到，
关于疑问1的想法：这种对比的时候应该是把第一个文档的每一行拎出来在第二个文档里全局查找
疑问2:涉及到符号和空格的怎么办？
关于疑问2的想法：一会儿实际操作遇上了再说
'''

#1对N对比

from  openpyxl import Workbook
from  openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

#读取sheet内容
name='//Users//zhangjingjing//desktop//diff2.xlsx'
wb=load_workbook(name)
ws=wb['Sheet1']
ws2=wb['Sheet2']

#1对N对比
#在sheet1中标出 sheet1与sheet2不同的地方

for row in ws[1:401]:
    for r in row:
        for row2 in ws2[1:401]:
            for r2 in row2:
                pass
            row_split1=(str(row).split('.'))[1]
            row_split2=(row_split1.split('>'))[0]
            #查找范围由sheet2的范围决定
            row2_split1=(str(row2).split('.'))[1]
            row2_split2=(row2_split1.split('>'))[0]
            if (str(r.value)).find(str(r2.value)) == -1 and row2_split2 == 'A401':
                print(row2_split2)
                ft = Font(color=colors.RED) 
                ws[row_split2].font=ft 
                print('已在第一个文档中标出不同的内容')
                print(r.value)
            elif (str(r.value)).find(str(r2.value)) != -1 : 
                print(row2_split2)
                print('已经找到了')
                print(r.value)
                break
            else :
                print(row2_split2)
                print('还未查找完毕')
                print(r.value)
                continue 
            

#N对1对比
#1对N对比
#在sheet2中标出 sheet2与sheet1不同的地方
for row2 in ws2[1:401]:
    for r2 in row2:
        for row in ws[1:401]:
            for r in row:
                pass
            row2_split1=(str(row2).split('.'))[1]
            row2_split2=(row2_split1.split('>'))[0]
            #查找范围由sheet1的范围决定
            row_split1=(str(row).split('.'))[1]
            row_split2=(row_split1.split('>'))[0]
            if (str(r2.value)).find(str(r.value)) == -1 and row_split2 == 'A401':
                print(row_split2)
                ft = Font(color=colors.GREEN) 
                ws2[row2_split2].font=ft 
                print('已在第二个文档中标出不同的内容')
                print(r2.value)
            elif (str(r2.value)).find(str(r.value)) != -1 : 
                print(row_split2)
                print('已经找到了')
                print(r2.value)
                break
            else :
                print(row_split2)
                print('还未查找完毕')
                print(r2.value)
                continue 

#保存
wb.save(name)
