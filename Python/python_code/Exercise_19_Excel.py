#coding=utf-8
#python对Excel进行操作
#参照博客https://www.cnblogs.com/zeke-python-road/p/8986318.html
from  openpyxl import Workbook
from  openpyxl import load_workbook
import time

#第一部分 工作簿与工作表
'''
name='python_excel_1.xlsx'
name2='python_excel_2.xlsx'
#1.新建Excel文件并保存;
#每个workbook创建后，会默认创建一个sheet
wb=Workbook()
wb.save(name)
print('creat successful')

#2.写入
ws=wb.active#print(ws)#<Worksheet "Sheet">
#2.1在第一个默认sheet中写入内容
ws['A1']='hello everybody it\'s my pleasure to show some of you that my diary about my life recently'
ws['A2']=time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())
#.2.2创建sheet
for i in range(1,3):
    ws=wb.create_sheet('Mysheet_'+ str(i))
    ws.title='Mysheet_'+ str(i)
    #ws.sheet_properties.tabcolor='1072ba'
    #wb.save('python_excel_'+ str(i)+'.xlsx')
wb.save(name)

#3.读取
#3.1读取工作表内的内容
wb2=load_workbook(name)
#<openpyxl.workbook.workbook.Workbook object at 0x10e92d890>
ws2=wb2.active #<Worksheet "Sheet">获取活动工作表
row=[]
for row in ws2.iter_rows():
    print(row[0].value)#hello everybody it's my pleasure to show some of you that my diary about my life recently
    print('\n')        #2022年03月06日 19时36分07秒

#3.2读取工作表的名称
#不同方式读取sheet的名称
#3.2.1
print('3.2.1')
for i in range(3):
    ws3=wb2.worksheets[i]#以索引方式获取工作表
    print(ws3)#<Worksheet "Sheet">#<Worksheet "Mysheet_1">#<Worksheet "Mysheet_2">
print('\n')   

#3.2.2 wb2.worksheets
print('3.2.2')
for i in wb2.worksheets:
    print(i)
print('\n')

#3.2.3 以工作表名获取
print('3.2.3')
ws4=wb['Mysheet_1']
print(ws4)#<Worksheet "Mysheet_1">
print('\n')

#3.2.4 sheetnames
print('3.2.4')
print(wb.sheetnames)
print('\n')

#3.2.5 title 获取/修改 工作表名
print('3.2.5') 
print(wb.worksheets[2].title)#Mysheet_2

wb2.save(name2) 



#4.批量新建工作簿/批量新建工作表
for i in range(3,6):
    wb3=Workbook()
    for j in range(1,4):
        ws5=wb3.create_sheet('Mysheet_'+ str(j))
        ws5.title='Mysheet_'+ str(j)
    wb3.save('python_excel_'+ str(i)+'.xlsx')

#5.批量修改工作表名 python_excel_3.xlsx
wb6=load_workbook('python_excel_3.xlsx')
print('#5')
for ws6 in wb6.worksheets:
    #print(ws6.title)
    ws6.title=ws6.title+'-add'
    print(ws6.title)
wb6.save('python_excel_3(1).xlsx')#相当于另存为
print('\n')


#6.复制/删除工作表
print('6.1')
wb7=load_workbook('python_excel_4.xlsx')
ws7=wb7.copy_worksheet(wb7['Sheet']).title='Sheet_copy'#括号里填写sheet对象
wb7.save('python_excel_4.xlsx')
print(ws7)#Sheet_copy
print('\n')

#6.2删除工作表
wb8=load_workbook('python_excel_5.xlsx')
wb8.remove(wb8['Sheet'])
wb8.save('python_excel_5.xlsx')


#7.批量复制/删除工作表
#7.1新建一个工作簿->在指定位置新建工作表
wb9name='python_excel_6.xlsx'
wb9=Workbook()
ws9=load_workbook(wb9name)
ws9=wb9.create_sheet('newsheet1',0)#create_sheet(指定sheet名称，指定sheet位置)
ws9=wb9.create_sheet('newsheet3')
ws9=wb9.create_sheet('newsheet2',1)
wb9.save(wb9name)

#7.1批量复制
wb10=load_workbook('python_excel_6.xlsx')
for ws10 in wb10.worksheets:
    if ws10.title.split('s')[0]== 'new' :
        wb10.copy_worksheet(ws10).title=ws10.title+'_copy'
wb10.save('python_excel_6.xlsx')

#7.2批量删除
wb11=load_workbook('python_excel_6.xlsx')
for ws11 in wb11.worksheets:
    if ws11.title.split('s')[0]== 'new' :
        wb11.remove(ws11)
wb11.save('python_excel_6.xlsx')



#第二部分 单元格数据
#1.单个单元格获取数据方法 ws['A1'],ws.cell(1,1)
wb12=load_workbook('python_excel_7.xlsx')
ws12=wb12.worksheets[0]
print(ws12['A1'].value)#hello
print(ws12.cell(1,1).value)#hello
print(load_workbook('python_excel_7.xlsx').worksheets[0]['A1'].value)#hello

#2 单元格区域数据获取
wb12=load_workbook('python_excel_7.xlsx')
ws12=wb12.active

#按行读取
#方法1
print(list(ws12.values)[1:5])

#方法2
print('\n') 
for row in ws12[1:14]:
    for r in row:
        if r.value!= None:
            print(r.value)     


#按列读取
print('\n') 
for col in ws12['a:g']:
    for c in col:
        if c.value!= None:
            print(c.value)


wb13=Workbook()
wb13.save('python_excel_8_sum.xlsx')
'''
wb14=load_workbook('python_excel_8_sum.xlsx')
ws14=wb14.active
count=0


for col in ws14['b:d']:
    for c in col:
        if c.value!= None:
            print(c.value)
            print(type(c.value))
            if type(c.value)==type(long()):
                count+=1
                sum=c.value
                print('sum1=', sum)
                print(type(sum))
    #print('sum2=', sum(c.value))











