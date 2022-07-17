#coding=utf-8
#txt 文档对比

import os
from  openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

last_name='lastWeekFiles\installed-files-boot.txt'
new_name='thisWeekFiles\installed-files-boot.txt'
result_name_1='result3.xlsx'


#文档
f1=open(last_name, mode='r+', encoding='utf-8')
f2=open(new_name, mode='r+', encoding='utf-8')

#结果输出文档
wb=Workbook()
ws=wb.active#print(ws)#<Worksheet "Sheet">
ws['A1']='上周Release文件'
ws['B1']='本周Release文件'
ws['C1']='上周Release文件大小'
ws['D1']='本周Release文件大小'
ws['E1']='上周Release与本周完全相同的文件'
ws['F1']='上周Release相对本周多出的文件'#颜色红色
ws['G1']='本周Release相对上周多出的文件'#颜色红色
ws['H1']='文件相同但大小有变化的文件'#颜色绿色
ws['I1']='变化量'
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 45
ws.column_dimensions['G'].width = 45
ws.column_dimensions['H'].width = 45
ws.column_dimensions['I'].width = 8
row1=row2=row3=row4=row5=row6=2

a=f1.readlines()
b=f2.readlines()
#1-4列填充
for last_line in a:
    ws.cell(row1, 1).value=last_line[14:]
    ws.cell(row1, 3).value=last_line[:14]
    row1=row1+1
for this_line in b:
    ws.cell(row2, 2).value=this_line[14:]
    ws.cell(row2, 4).value=this_line[:14]
    row2=row2+1

#填充第5列：完全相同文件； 填充数据不同，文件相同
for last_row,last_data in zip(ws['A'],ws['C']):
    for this_row,this_data in zip(ws['B'],ws['D']):
        if last_row.value == this_row.value and last_data.value == this_data.value:
            print('11111111111111111111111111111111')
            print('last_row.value: '+ last_row.value)
            print('this_row.value: '+ this_row.value)
            print('last_data.value: '+ last_data.value)
            print('last_data.value: '+ this_data.value)
            ws.cell(row3,5).value=last_row.value
            row3=row3+1
        elif last_row.value == this_row.value and last_data.value != this_data.value:
            print('22222222222222222222222222222222')
            print('last_row.value: '+ last_row.value)
            print('this_row.value: '+ this_row.value)
            print('last_data.value: '+ last_data.value)
            print('last_data.value: '+ this_data.value)
            ws.cell(row4,8).value=last_row.value
            ft = Font(color=colors.RED)
            ws.cell(row4,8).font=ft  
            ws.cell(row4,9).value=str(int(this_data.value)-int(last_data.value))
            row4=row4+1


for last_row, this_row  in zip(range(2, len(ws['A'])) ,range(2,len(ws['B']))) :
    for com_row in range(2, len(ws['E']) ) :
        #common文件
        if ws.cell(last_row, 1).value == ws.cell(com_row,5).value  or ws.cell(this_row,2).value ==  ws.cell(com_row,5).value :
            break
        elif ws.cell(last_row, 1).value != ws.cell(com_row,5).value and com_row==414:
            print('33333333333333333333333333333333-#上 有 本 无')
            print('last_row.value: '+ ws.cell(last_row, 1).value)
            print('com_row.value: '+  ws.cell(com_row,5).value)
            ws.cell(row5,6).value=ws.cell(last_row, 1).value
        elif ws.cell(this_row,2).value !=  ws.cell(com_row,5).value and com_row==414:
            print('44444444444444444444444444444444-#上 无 本 有')
            print('this_row.value: '+ ws.cell(this_row,2).value )
            print('com_row.value: '+  ws.cell(com_row,5).value)
            ws.cell(row6,7).value=ws.cell(this_row,2).value    
        

wb.save(result_name_1)
f1.close()
f2.close()