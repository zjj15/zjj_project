#coding=utf-8
#txt 文档对比

import os
from  openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

last_name='lastWeekFiles\installed-files-boot.txt'
new_name='thisWeekFiles\installed-files-boot.txt'
result_name_1='result.xlsx'


#文档
f1=open(last_name, mode='r+', encoding='utf-8')
f2=open(new_name, mode='r+', encoding='utf-8')

#结果输出文档
wb=Workbook()
ws=wb.active#print(ws)#<Worksheet "Sheet">
ws['A1']='上周Release相对本周多出的文件'
ws['B1']='本周Release相对上周多出的文件'
ws['C1']='大小无变化的文件'
ws['D1']='大小有变化的文件'
ws['E1']='变化量'
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 8
row1=row2=row3=row4=row5=row6=2

a=f1.readlines()
b=f2.readlines()
#上周与本周对比
for f1_readline in a:
    for f2_readline  in b:
        #1.文件+数量完全相同
        if f2_readline[:] == f1_readline[:] :
            print('1.大小无变化的文件 '+ f1_readline)
            ws.cell(row1,3).value=f1_readline[14:]
            row1=row1+1 
            break
        #2.文件相同，数量不同
        elif  f1_readline[14:] == f2_readline[14:] and f1_readline[0:14] != f2_readline[0:14]:
        #f1_readline[0:f1_readline.find('  /')] != f2_readline[0:f2_readline.find('  /')]   and  f1_readline[f1_readline.find('  /'):] == f2_readline[f2_readline.find('  /'):] :
            print('2.大小有变化的文件' +f1_readline)
            ws.cell(row2,4).value=f1_readline[14:]
            ft = Font(color=colors.RED)
            ws.cell(row2,4).font=ft  
        
            print('2.1变化量')
            ws.cell(row2,5).value=str(int(f1_readline[0:14] )- int(f2_readline[0:14]))
            row2=row2+1
            break
        #3.文件不同，未比到最后一个
        elif f1_readline[14:] !=f2_readline[14:] and f2_readline != b[-1]:
            print(f1_readline[14:]!=f2_readline[14:])
            print('f1_readline[14:]: '+ f1_readline)
            print('f2_readline[14:]: '+ f2_readline)
            print('b[-1]: '+ b[-1])
            print('3.文件不同且不是最后一行，继续判断下一行f2'+ f2_readline)
        #4.文件不同，已比到最后一个
        elif f1_readline[14:] != f2_readline[14:] and f2_readline == b[-1]: 
            print('f1_readline: '+ f1_readline)
            print('4.文件不同且是最后一行f2: '+ f2_readline)
            print('4.1上周Release相对本周多出的文件'+f1_readline)
            ws.cell(row3,1).value=f1_readline[14:]
            row3=row3+1
        else:
            print('error '+f1_readline[14:] )
            

#本周与上周对比
for f2_readline in b:
    for f1_readline  in a:
        
        #1.文件+数量完全相同
        if f1_readline[:] == f2_readline[:] :
            print('1.大小无变化的文件 '+ f2_readline)
            #ws.cell(row1,3).value=f2_readline[14:]
            row1=row1+1 
            break
            '''
        #2.文件相同，数量不同(只需要1次)
        elif  f2_readline[14:] == f1_readline[14:] and f2_readline[0:14] != f1_readline[0:14]:
        #f2_readline[0:f2_readline.find('  /')] != f1_readline[0:f1_readline.find('  /')]   and  f2_readline[f2_readline.find('  /'):] == f1_readline[f1_readline.find('  /'):] :
            print('2.大小有变化的文件' +f2_readline)
            ws.cell(row2,4).value=f2_readline[14:]
            ft = Font(color=colors.RED)
            ws.cell(row2,4).font=ft  
        
            print('2.1变化量')
            ws.cell(row2,5).value=str(int(f2_readline[0:14] )- int(f1_readline[0:14]))
            row2=row2+1
            break
            '''
        #3.文件不同，未比到最后一个
        elif f2_readline[14:] !=f1_readline[14:] and f1_readline != a[-1]:
            print(f2_readline[14:]!=f1_readline[14:])
            print('f2_readline[14:]: '+ f2_readline)
            print('f1_readline[14:]: '+ f1_readline)
            print('b[-1]: '+ b[-1])
            print('3.文件不同且不是最后一行，继续判断下一行f1'+ f1_readline)
        #4.文件不同，已比到最后一个
        elif f2_readline[14:] != f1_readline[14:] and f1_readline == a[-1]: 
            print('f2_readline: '+ f2_readline)
            print('4.文件不同且是最后一行f1: '+ f1_readline)
            print('4.1本周Release相对上周多出的文件'+f2_readline)
            ws.cell(row3,2).value=f2_readline[14:]
            row3=row3+1
                  

wb.save(result_name_1)
f1.close()
f2.close()
